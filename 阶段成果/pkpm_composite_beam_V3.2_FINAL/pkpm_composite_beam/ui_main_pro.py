"""
PKPM-CAE 叠合梁参数化建模引擎 - PyQt5 专业版UI
T+7 优化版 - 印刷级界面质量
"""

import sys
import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QGroupBox, QFormLayout, QTextEdit, QMessageBox, QScrollArea,
    QDoubleSpinBox, QSpinBox, QComboBox, QFrame
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor, QPixmap, QPainter

# 核心路径修复逻辑
def get_resource_path(relative_path):
    """获取程序运行时资源的绝对路径（兼容源码和EXE打包）"""
    if hasattr(sys, '_MEIPASS'):
        # EXE 运行时，指向临时目录
        return os.path.join(sys._MEIPASS, relative_path)
    # 源码运行时，指向当前目录
    return os.path.join(os.path.abspath("."), relative_path)

# 修改所有涉及路径的地方
current_dir = get_resource_path("")
sys.path.insert(0, current_dir)

try:
    from main import CompositeBeamModelGenerator
    ENGINE_AVAILABLE = True
except Exception as e:
    print(f"警告: 主引擎模块未加载 - {e}")
    ENGINE_AVAILABLE = False


class ModelGenerationThread(QThread):
    """后台模型生成线程"""
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, excel_path, output_script="pkpm_composite_beam_model.py"):
        super().__init__()
        self.excel_path = excel_path
        self.output_script = output_script

    def run(self):
        try:
            self.progress.emit("[启动] 初始化引擎...")
            generator = CompositeBeamModelGenerator(self.excel_path)

            self.progress.emit("[1/7] 解析 Excel 参数...")
            generator.parse_excel()

            self.progress.emit("[2/7] 创建几何模型...")
            generator.create_geometry()

            self.progress.emit("[3/7] 创建钢筋布置...")
            generator.create_rebars()

            self.progress.emit("[4/7] 配置钢筋嵌入...")
            generator.create_embedment()

            self.progress.emit("[5/7] 创建预应力孔道...")
            generator.create_prestress_ducts()

            self.progress.emit("[6/7] 配置两阶段分析...")
            generator.create_two_stage_analysis()

            self.progress.emit("[7/7] 导出 Python 脚本...")
            generator.export_script(self.output_script)

            self.finished.emit(True, f"✅ 模型生成成功！\n输出文件: {self.output_script}")
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.finished.emit(False, f"❌ 错误: {str(e)}\n\n详细信息:\n{error_detail}")


class CompositeBeamUI(QMainWindow):
    """PKPM-CAE 叠合梁参数化建模 专业版UI"""

    def __init__(self):
        super().__init__()
        self.excel_path = None
        self.init_ui()
        self.load_demo_parameters()  # 自动加载演示参数

    @staticmethod
    def create_label(text):
        """创建带中文字体的标签"""
        label = QLabel(text)
        label.setFont(QFont("Microsoft YaHei", 10))
        return label

    def init_ui(self):
        """初始化专业级UI界面"""
        self.setWindowTitle("PKPM-CAE 叠合梁参数化建模引擎 v1.0 (T+7)")
        self.setGeometry(100, 100, 1200, 850)

        # 设置全局默认中文字体
        default_font = QFont("Microsoft YaHei", 10)
        QApplication.instance().setFont(default_font)

        # 设置应用程序样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F3F4F6;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #D1D5DB;
                border-radius: 8px;
                margin-top: 12px;
                padding: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #1F2937;
            }
            QLabel {
                color: #374151;
            }
            QLineEdit, QDoubleSpinBox, QSpinBox, QComboBox {
                padding: 6px;
                border: 1px solid #D1D5DB;
                border-radius: 4px;
                background-color: white;
                min-height: 25px;
            }
            QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus {
                border: 2px solid #3B82F6;
            }
            QTabWidget::pane {
                border: 2px solid #E5E7EB;
                border-radius: 6px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #E5E7EB;
                color: #4B5563;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #3B82F6;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #60A5FA;
                color: white;
            }
        """)

        # 中央窗口部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # ========== Logo 和标题区 ==========
        header_widget = self.create_header()
        main_layout.addWidget(header_widget)

        # ========== Excel 文件选择区 ==========
        file_group = QGroupBox("📁 Excel 参数文件")
        file_layout = QHBoxLayout()
        file_group.setLayout(file_layout)

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("请选择 Excel 参数文件（或使用默认演示参数）")
        self.file_path_edit.setReadOnly(True)
        self.file_path_edit.setStyleSheet("font-size: 13px;")
        file_layout.addWidget(self.file_path_edit, 3)

        browse_btn = QPushButton("📂 浏览...")
        browse_btn.clicked.connect(self.browse_excel)
        browse_btn.setStyleSheet("""
            QPushButton {
                padding: 8px 20px;
                font-weight: bold;
                background-color: #6B7280;
                color: white;
                border-radius: 4px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #4B5563;
            }
        """)
        file_layout.addWidget(browse_btn)

        load_btn = QPushButton("📥 读取 Excel")
        load_btn.clicked.connect(self.load_excel)
        load_btn.setStyleSheet("""
            QPushButton {
                padding: 8px 20px;
                font-weight: bold;
                background-color: #3B82F6;
                color: white;
                border-radius: 4px;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #2563EB;
            }
        """)
        file_layout.addWidget(load_btn)

        main_layout.addWidget(file_group)

        # ========== 参数输入区（6 个标签页）==========
        self.tab_widget = QTabWidget()
        self.tab_widget.setFont(QFont("Microsoft YaHei", 11))  # 设置标签页字体
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #3B82F6;
                border-radius: 8px;
                background-color: white;
                padding: 10px;
            }
            QTabBar::tab {
                font-family: "Microsoft YaHei";
                font-size: 13px;
                font-weight: bold;
            }
        """)

        # Tab 1: 几何参数
        self.create_geometry_tab()

        # Tab 2: 纵向配筋
        self.create_rebar_tab()

        # Tab 3: 箍筋
        self.create_stirrup_tab()

        # Tab 4: 洞口与倒角
        self.create_holes_tab()

        # Tab 5: 荷载与边界
        self.create_loads_tab()

        # Tab 6: 预应力
        self.create_prestress_tab()

        main_layout.addWidget(self.tab_widget, 1)

        # ========== 操作按钮区 ==========
        btn_layout = QHBoxLayout()

        generate_btn = QPushButton("🚀 一键生成模型")
        generate_btn.clicked.connect(self.generate_model)
        generate_btn.setStyleSheet("""
            QPushButton {
                padding: 15px 40px;
                font-size: 18px;
                font-weight: bold;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #10B981, stop:1 #059669);
                color: white;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #059669, stop:1 #047857);
            }
            QPushButton:pressed {
                background: #047857;
            }
        """)
        btn_layout.addStretch()
        btn_layout.addWidget(generate_btn)
        btn_layout.addStretch()

        main_layout.addLayout(btn_layout)

        # ========== 日志输出区 ==========
        log_group = QGroupBox("📊 生成日志")
        log_layout = QVBoxLayout()
        log_group.setLayout(log_layout)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(120)
        self.log_text.setStyleSheet("""
            background-color: #1F2937;
            color: #10B981;
            font-family: 'Consolas', 'Courier New', monospace;
            font-size: 12px;
            padding: 10px;
            border-radius: 4px;
        """)
        self.log_text.append(">>> 系统已就绪，等待参数输入...")
        self.log_text.append(">>> 已自动加载演示参数 (10m 跨叠合梁)")
        log_layout.addWidget(self.log_text)

        main_layout.addWidget(log_group)

    def create_header(self):
        """创建Logo和标题区"""
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                            stop:0 #1E40AF, stop:1 #3B82F6);
                border-radius: 10px;
                padding: 20px;
            }
        """)

        header_layout = QVBoxLayout()
        header_frame.setLayout(header_layout)

        # Logo 文字
        logo_label = QLabel("PKPM-CAE")
        logo_label.setAlignment(Qt.AlignCenter)
        logo_font = QFont("Arial", 24, QFont.Bold)
        logo_label.setFont(logo_font)
        logo_label.setStyleSheet("color: white; padding: 5px;")
        header_layout.addWidget(logo_label)

        # 主标题
        title_label = QLabel("叠合梁参数化建模自动化工具")
        title_label.setAlignment(Qt.AlignCenter)
        title_font = QFont("Microsoft YaHei", 20, QFont.Bold)
        title_label.setFont(title_font)
        title_label.setStyleSheet("color: white; padding: 5px;")
        header_layout.addWidget(title_label)

        # 副标题
        subtitle_label = QLabel("T+7 专业版 | Excel驱动 • 一键生成 • 工程级质量")
        subtitle_label.setAlignment(Qt.AlignCenter)
        subtitle_font = QFont("Microsoft YaHei", 11)
        subtitle_label.setFont(subtitle_font)
        subtitle_label.setStyleSheet("color: #E0E7FF; padding: 3px;")
        header_layout.addWidget(subtitle_label)

        return header_frame

    def create_geometry_tab(self):
        """创建几何参数标签页"""
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(tab)
        scroll.setWidgetResizable(True)

        main_layout = QVBoxLayout()
        tab.setLayout(main_layout)

        self.geom_inputs = {}

        # ========== 截面类型选择组 ==========
        section_group = QGroupBox("截面类型选择")
        section_layout = QFormLayout()
        section_group.setLayout(section_layout)

        section_type_combo = QComboBox()
        section_type_combo.addItems(["矩形截面", "T型截面", "倒T型截面 (常用)", "工字型截面"])
        section_type_combo.setCurrentIndex(2)  # 默认倒T型
        section_type_combo.currentIndexChanged.connect(self._on_section_type_changed)
        self.geom_inputs["section_type"] = section_type_combo

        label = QLabel("截面形式:")
        label.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        label.setStyleSheet("color: #1F2937;")
        section_layout.addRow(label, section_type_combo)

        # 截面示意图说明
        section_info = QLabel("矩形=无翼缘 | T型=上翼缘 | 倒T型=下翼缘 | 工字型=上下翼缘")
        section_info.setStyleSheet("color: #6B7280; font-size: 11px; font-style: italic;")
        section_layout.addRow(section_info)

        main_layout.addWidget(section_group)

        # ========== 基本尺寸组 ==========
        basic_group = QGroupBox("基本尺寸")
        basic_layout = QFormLayout()
        basic_group.setLayout(basic_layout)

        basic_fields = [
            ("L", "梁长", 10000.0, "mm"),
            ("H", "梁高", 800.0, "mm"),
            ("Tw", "腹板宽度", 250.0, "mm"),
            ("h_pre", "预制层高度", 500.0, "mm"),
            ("t_cast_cap", "现浇顶盖厚度(0=自动)", 75.0, "mm"),
        ]

        for field_name, label_text, default, unit in basic_fields:
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 100000)
            input_widget.setValue(default)
            input_widget.setDecimals(1)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))
            label.setStyleSheet("font-weight: bold; color: #374151;")

            self.geom_inputs[field_name] = input_widget
            basic_layout.addRow(label, input_widget)

        main_layout.addWidget(basic_group)

        # ========== 上翼缘参数组 ==========
        self.upper_flange_group = QGroupBox("上翼缘参数")
        upper_flange_layout = QFormLayout()
        self.upper_flange_group.setLayout(upper_flange_layout)

        upper_flange_fields = [
            ("bf_lu", "左上翼缘伸出宽", 100.0),
            ("tf_lu", "左上翼缘厚度", 150.0),
            ("bf_ru", "右上翼缘伸出宽", 100.0),
            ("tf_ru", "右上翼缘厚度", 150.0),
        ]

        for field_name, label_text, default in upper_flange_fields:
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 5000)
            input_widget.setValue(default)
            input_widget.setDecimals(1)
            input_widget.setSuffix(" mm")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))
            label.setStyleSheet("color: #374151;")

            self.geom_inputs[field_name] = input_widget
            upper_flange_layout.addRow(label, input_widget)

        main_layout.addWidget(self.upper_flange_group)

        # ========== 下翼缘参数组 ==========
        self.lower_flange_group = QGroupBox("下翼缘参数")
        lower_flange_layout = QFormLayout()
        self.lower_flange_group.setLayout(lower_flange_layout)

        lower_flange_fields = [
            ("bf_ll", "左下翼缘伸出宽", 100.0),
            ("tf_ll", "左下翼缘厚度", 150.0),
            ("bf_rl", "右下翼缘伸出宽", 100.0),
            ("tf_rl", "右下翼缘厚度", 150.0),
        ]

        for field_name, label_text, default in lower_flange_fields:
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 5000)
            input_widget.setValue(default)
            input_widget.setDecimals(1)
            input_widget.setSuffix(" mm")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))
            label.setStyleSheet("color: #374151;")

            self.geom_inputs[field_name] = input_widget
            lower_flange_layout.addRow(label, input_widget)

        main_layout.addWidget(self.lower_flange_group)

        # ========== 混凝土等级组 ==========
        concrete_group = QGroupBox("混凝土等级")
        concrete_layout = QFormLayout()
        concrete_group.setLayout(concrete_layout)

        precast_combo = QComboBox()
        precast_combo.addItems(["C30", "C35", "C40", "C45", "C50"])
        precast_combo.setCurrentText("C40")
        self.geom_inputs["precast_concrete_grade"] = precast_combo

        cast_combo = QComboBox()
        cast_combo.addItems(["C30", "C35", "C40", "C45", "C50"])
        cast_combo.setCurrentText("C35")
        self.geom_inputs["cast_concrete_grade"] = cast_combo

        concrete_layout.addRow(self.create_label("预制层等级:"), precast_combo)
        concrete_layout.addRow(self.create_label("现浇层等级:"), cast_combo)

        main_layout.addWidget(concrete_group)
        main_layout.addStretch()

        self.tab_widget.addTab(scroll, "📐 几何参数")

        # 初始化时根据默认选择更新UI状态
        self._on_section_type_changed(2)  # 默认倒T型

    def _on_section_type_changed(self, index):
        """截面类型切换时更新翼缘参数的启用状态"""
        # index: 0=矩形, 1=T型, 2=倒T型, 3=工字型

        # 上翼缘参数列表
        upper_params = ['bf_lu', 'tf_lu', 'bf_ru', 'tf_ru']
        # 下翼缘参数列表
        lower_params = ['bf_ll', 'tf_ll', 'bf_rl', 'tf_rl']
        # 叠合面切分：现浇顶盖厚度仅在有上翼缘时有效
        has_upper_flange = index in (1, 3)
        if "t_cast_cap" in self.geom_inputs:
            self.geom_inputs["t_cast_cap"].setEnabled(has_upper_flange)
            if not has_upper_flange:
                self.geom_inputs["t_cast_cap"].setValue(0.0)

        if index == 0:  # 矩形截面
            # 禁用所有翼缘，设为0
            self.upper_flange_group.setEnabled(False)
            self.lower_flange_group.setEnabled(False)
            for p in upper_params + lower_params:
                self.geom_inputs[p].setValue(0)
            self.upper_flange_group.setTitle("上翼缘参数 (矩形截面不需要)")
            self.lower_flange_group.setTitle("下翼缘参数 (矩形截面不需要)")

        elif index == 1:  # T型截面
            # 启用上翼缘，禁用下翼缘
            self.upper_flange_group.setEnabled(True)
            self.lower_flange_group.setEnabled(False)
            for p in upper_params:
                if self.geom_inputs[p].value() == 0:
                    self.geom_inputs[p].setValue(100.0 if 'bf' in p else 150.0)
            for p in lower_params:
                self.geom_inputs[p].setValue(0)
            self.upper_flange_group.setTitle("上翼缘参数 ✓")
            self.lower_flange_group.setTitle("下翼缘参数 (T型截面不需要)")

        elif index == 2:  # 倒T型截面
            # 禁用上翼缘，启用下翼缘
            self.upper_flange_group.setEnabled(False)
            self.lower_flange_group.setEnabled(True)
            for p in upper_params:
                self.geom_inputs[p].setValue(0)
            for p in lower_params:
                if self.geom_inputs[p].value() == 0:
                    self.geom_inputs[p].setValue(100.0 if 'bf' in p else 150.0)
            self.upper_flange_group.setTitle("上翼缘参数 (倒T型截面不需要)")
            self.lower_flange_group.setTitle("下翼缘参数 ✓")

        elif index == 3:  # 工字型截面
            # 启用所有翼缘
            self.upper_flange_group.setEnabled(True)
            self.lower_flange_group.setEnabled(True)
            for p in upper_params + lower_params:
                if self.geom_inputs[p].value() == 0:
                    self.geom_inputs[p].setValue(100.0 if 'bf' in p else 150.0)
            self.upper_flange_group.setTitle("上翼缘参数 ✓")
            self.lower_flange_group.setTitle("下翼缘参数 ✓")

    def create_rebar_tab(self):
        """创建纵向配筋标签页"""
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(tab)
        scroll.setWidgetResizable(True)

        main_layout = QVBoxLayout()
        tab.setLayout(main_layout)

        self.rebar_inputs = {}

        # 顶部钢筋组
        top_group = QGroupBox("顶部纵向钢筋")
        top_layout = QFormLayout()
        top_group.setLayout(top_layout)

        top_fields = [
            ("top_dia", "钢筋直径", 20, "mm"),
            ("top_num", "钢筋根数", 4, "根"),
            ("top_spacing", "横向间距", 80, "mm"),
            ("top_cover", "保护层厚度", 40, "mm"),
        ]

        for field_name, label_text, default, unit in top_fields:
            input_widget = QSpinBox()
            input_widget.setRange(0, 1000)
            input_widget.setValue(default)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))  # 明确设置中文字体
            label.setStyleSheet("font-weight: bold; color: #374151;")

            self.rebar_inputs[field_name] = input_widget
            top_layout.addRow(label, input_widget)

        main_layout.addWidget(top_group)

        # 底部钢筋组
        bottom_group = QGroupBox("底部纵向钢筋")
        bottom_layout = QFormLayout()
        bottom_group.setLayout(bottom_layout)

        bottom_fields = [
            ("bottom_dia", "钢筋直径", 25, "mm"),
            ("bottom_num", "钢筋根数", 6, "根"),
            ("bottom_spacing", "横向间距", 70, "mm"),
            ("bottom_cover", "保护层厚度", 40, "mm"),
        ]

        for field_name, label_text, default, unit in bottom_fields:
            input_widget = QSpinBox()
            input_widget.setRange(0, 1000)
            input_widget.setValue(default)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))  # 明确设置中文字体
            label.setStyleSheet("font-weight: bold; color: #374151;")

            self.rebar_inputs[field_name] = input_widget
            bottom_layout.addRow(label, input_widget)

        main_layout.addWidget(bottom_group)
        main_layout.addStretch()

        self.tab_widget.addTab(scroll, "🔩 纵向配筋")

    def create_stirrup_tab(self):
        """创建箍筋标签页"""
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(tab)
        scroll.setWidgetResizable(True)

        main_layout = QVBoxLayout()
        tab.setLayout(main_layout)

        self.stirrup_inputs = {}

        group = QGroupBox("箍筋配置")
        layout = QFormLayout()
        group.setLayout(layout)

        fields = [
            ("stirrup_dia", "箍筋直径", 10, "mm"),
            ("stirrup_dense_spacing", "加密区间距", 100, "mm"),
            ("stirrup_normal_spacing", "非加密区间距", 200, "mm"),
            ("stirrup_dense_length", "端部加密长度", 1500, "mm"),
            ("stirrup_legs", "箍筋肢数", 4, "肢"),
            ("stirrup_cover", "箍筋保护层", 25, "mm"),
        ]

        for field_name, label_text, default, unit in fields:
            input_widget = QSpinBox()
            input_widget.setRange(0, 5000)
            input_widget.setValue(default)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))  # 明确设置中文字体
            label.setStyleSheet("font-weight: bold; color: #374151;")

            self.stirrup_inputs[field_name] = input_widget
            layout.addRow(label, input_widget)

        main_layout.addWidget(group)
        main_layout.addStretch()

        self.tab_widget.addTab(scroll, "⚙️ 箍筋")

    def create_holes_tab(self):
        """创建洞口与倒角标签页"""
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(tab)
        scroll.setWidgetResizable(True)

        main_layout = QVBoxLayout()
        tab.setLayout(main_layout)

        self.hole_inputs = {}

        # 洞口1组
        hole1_group = QGroupBox("洞口 1")
        hole1_layout = QFormLayout()
        hole1_group.setLayout(hole1_layout)

        hole1_fields = [
            ("hole1_x", "距左端距离", 2000, "mm"),
            ("hole1_z", "距底部距离", 100, "mm"),
            ("hole1_width", "洞口宽度", 800, "mm"),
            ("hole1_height", "洞口高度", 300, "mm"),
        ]

        for field_name, label_text, default, unit in hole1_fields:
            input_widget = QSpinBox()
            input_widget.setRange(0, 20000)
            input_widget.setValue(default)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))  # 明确设置中文字体
            label.setStyleSheet("font-weight: bold; color: #374151;")

            self.hole_inputs[field_name] = input_widget
            hole1_layout.addRow(label, input_widget)

        main_layout.addWidget(hole1_group)

        # 小梁配筋组 (洞口上下的小梁)
        small_beam_group = QGroupBox("小梁配筋 (洞口上下)")
        small_beam_layout = QFormLayout()
        small_beam_group.setLayout(small_beam_layout)

        small_beam_fields = [
            ("smallbeam_long_dia", "纵筋直径", 16, "mm"),
            ("smallbeam_long_count", "纵筋根数", 2, "根"),
            ("smallbeam_stirrup_dia", "箍筋直径", 8, "mm"),
            ("smallbeam_stirrup_spacing", "箍筋间距", 150, "mm"),
        ]

        for field_name, label_text, default, unit in small_beam_fields:
            input_widget = QSpinBox()
            input_widget.setRange(0, 1000)
            input_widget.setValue(default)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))
            label.setStyleSheet("color: #374151;")

            self.hole_inputs[field_name] = input_widget
            small_beam_layout.addRow(label, input_widget)

        main_layout.addWidget(small_beam_group)

        # 侧边补强组
        side_reinf_group = QGroupBox("侧边补强")
        side_reinf_layout = QFormLayout()
        side_reinf_group.setLayout(side_reinf_layout)

        side_reinf_fields = [
            ("left_reinf_length", "左侧补强长度", 500, "mm"),
            ("right_reinf_length", "右侧补强长度", 500, "mm"),
            ("side_stirrup_spacing", "侧边箍筋间距", 100, "mm"),
            ("side_stirrup_dia", "侧边箍筋直径", 10, "mm"),
            ("side_stirrup_legs", "侧边箍筋肢数", 2, "肢"),
            ("reinf_extend_length", "补强筋伸出长度", 300, "mm"),
        ]

        for field_name, label_text, default, unit in side_reinf_fields:
            input_widget = QSpinBox()
            input_widget.setRange(0, 5000)
            input_widget.setValue(default)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))
            label.setStyleSheet("color: #374151;")

            self.hole_inputs[field_name] = input_widget
            side_reinf_layout.addRow(label, input_widget)

        main_layout.addWidget(side_reinf_group)

        # 圆弧倒角组（T+7新功能）
        fillet_group = QGroupBox("圆弧倒角设置 (T+7 新功能)")
        fillet_layout = QFormLayout()
        fillet_group.setLayout(fillet_layout)

        fillet_enable = QComboBox()
        fillet_enable.addItems(["禁用", "启用"])
        fillet_enable.setCurrentText("启用")
        self.hole_inputs["fillet_enabled"] = fillet_enable

        fillet_radius = QDoubleSpinBox()
        fillet_radius.setRange(0, 500)
        fillet_radius.setValue(50.0)
        fillet_radius.setDecimals(1)
        fillet_radius.setSuffix(" mm")
        fillet_radius.setMinimumWidth(150)
        self.hole_inputs["fillet_radius"] = fillet_radius

        fillet_layout.addRow(self.create_label("倒角启用:"), fillet_enable)
        fillet_layout.addRow(self.create_label("倒角半径:"), fillet_radius)

        info_label = self.create_label("说明: 对洞口四角进行圆弧倒角，使几何更贴近实际工程")
        info_label.setStyleSheet("color: #6B7280; font-size: 11px; font-style: italic;")
        fillet_layout.addRow(info_label)

        main_layout.addWidget(fillet_group)
        main_layout.addStretch()

        self.tab_widget.addTab(scroll, "🔲 洞口 & 倒角")

    def create_loads_tab(self):
        """创建荷载与边界标签页"""
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(tab)
        scroll.setWidgetResizable(True)

        main_layout = QVBoxLayout()
        tab.setLayout(main_layout)

        self.load_inputs = {}

        # 荷载配置组
        load_group = QGroupBox("荷载配置")
        load_layout = QFormLayout()
        load_group.setLayout(load_layout)

        load_fields = [
            ("dead_load", "恒载 (自重+装修)", 15.0, "kN/m"),
            ("live_load", "活载", 20.0, "kN/m"),
        ]

        for field_name, label_text, default, unit in load_fields:
            input_widget = QDoubleSpinBox()
            input_widget.setRange(0, 1000)
            input_widget.setValue(default)
            input_widget.setDecimals(1)
            input_widget.setSuffix(f" {unit}")
            input_widget.setMinimumWidth(150)

            label = QLabel(f"{label_text}:")
            label.setFont(QFont("Microsoft YaHei", 10))
            label.setStyleSheet("font-weight: bold; color: #374151;")

            self.load_inputs[field_name] = input_widget
            load_layout.addRow(label, input_widget)

        main_layout.addWidget(load_group)

        # 边界条件组
        boundary_group = QGroupBox("边界条件")
        boundary_layout = QFormLayout()
        boundary_group.setLayout(boundary_layout)

        boundary_combo = QComboBox()
        boundary_combo.addItems(["一端固支一端简支 (推荐)", "两端简支", "两端固支"])
        boundary_combo.setCurrentIndex(0)
        self.load_inputs["boundary_condition"] = boundary_combo

        label = QLabel("支座类型:")
        label.setFont(QFont("Microsoft YaHei", 10))
        label.setStyleSheet("font-weight: bold; color: #374151;")
        boundary_layout.addRow(label, boundary_combo)

        # 边界说明
        info_label = self.create_label("说明: 左端为固定支座(Coupling刚性耦合), 右端为简支(线约束Dof.Uz)")
        info_label.setStyleSheet("color: #059669; font-size: 11px; font-style: italic;")
        info_label.setWordWrap(True)
        boundary_layout.addRow(info_label)

        main_layout.addWidget(boundary_group)

        # 荷载工况组
        case_group = QGroupBox("荷载工况")
        case_layout = QFormLayout()
        case_group.setLayout(case_layout)

        case_combo = QComboBox()
        case_combo.addItems(["标准组合", "准永久组合", "基本组合"])
        case_combo.setCurrentText("标准组合")
        self.load_inputs["load_case"] = case_combo

        label = QLabel("组合类型:")
        label.setFont(QFont("Microsoft YaHei", 10))
        label.setStyleSheet("font-weight: bold; color: #374151;")
        case_layout.addRow(label, case_combo)

        main_layout.addWidget(case_group)
        main_layout.addStretch()

        self.tab_widget.addTab(scroll, "📊 荷载 & 边界")

    def create_prestress_tab(self):
        """创建预应力标签页"""
        tab = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(tab)
        scroll.setWidgetResizable(True)

        main_layout = QVBoxLayout()
        tab.setLayout(main_layout)

        self.prestress_inputs = {}

        group = QGroupBox("预应力参数 (后台逻辑已完成)")
        layout = QFormLayout()
        group.setLayout(layout)

        prestress_enable = QComboBox()
        prestress_enable.addItems(["禁用", "启用"])
        prestress_enable.setCurrentText("禁用")
        self.prestress_inputs["enabled"] = prestress_enable

        # 注意：生成脚本侧 PreStress.value 以“应力(MPa)”处理（不是力N）
        prestress_force = QDoubleSpinBox()
        prestress_force.setRange(0, 3000)
        prestress_force.setValue(0)
        prestress_force.setDecimals(1)
        prestress_force.setSuffix(" MPa")
        prestress_force.setMinimumWidth(150)
        self.prestress_inputs["force"] = prestress_force

        prestress_method = QComboBox()
        prestress_method.addItems(["后张法(post_tension)", "先张法(pretension)"])
        prestress_method.setCurrentText("后张法(post_tension)")
        self.prestress_inputs["method"] = prestress_method

        duct_dia = QDoubleSpinBox()
        duct_dia.setRange(0, 200)
        duct_dia.setValue(90)
        duct_dia.setDecimals(1)
        duct_dia.setSuffix(" mm")
        duct_dia.setMinimumWidth(150)
        self.prestress_inputs["duct_diameter"] = duct_dia

        layout.addRow(self.create_label("预应力启用:"), prestress_enable)
        layout.addRow(self.create_label("张拉力:"), prestress_force)
        layout.addRow(self.create_label("预应力方式:"), prestress_method)
        layout.addRow(self.create_label("波纹管直径:"), duct_dia)

        info_label = self.create_label(
            "说明:\n"
            "- 后张法：预留波纹管孔道(duct_diameter>0)，并在分析阶段施加预应力。\n"
            "- 先张法：不挖孔道，直接对预应力筋/钢筋施加预应力（更利于网格稳定）。"
        )
        info_label.setStyleSheet("color: #059669; font-size: 11px; font-style: italic; padding: 10px;")
        info_label.setWordWrap(True)
        layout.addRow(info_label)

        main_layout.addWidget(group)
        main_layout.addStretch()

        self.tab_widget.addTab(scroll, "⚡ 预应力")

    def load_demo_parameters(self):
        """自动加载演示参数"""
        # 几何参数已在创建时设置默认值
        # 这里可以添加额外的演示数据加载逻辑
        self.file_path_edit.setText("【演示模式】使用默认参数 (10m跨叠合梁，C40预制+C35现浇)")
        pass

    def browse_excel(self):
        """浏览Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 参数文件", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.excel_path = file_path
            self.file_path_edit.setText(file_path)
            self.log_text.append(f">>> 已选择文件: {Path(file_path).name}")

    # def load_excel(self):
        """读取Excel文件"""
        if not self.excel_path:
            QMessageBox.warning(self, "警告", "请先选择 Excel 文件")
            return

        try:
            self.log_text.append(f">>> 正在解析 Excel: {Path(self.excel_path).name}")
            # 这里添加 Excel 读取逻辑
            self.log_text.append(">>> Excel 解析完成，参数已加载到界面")
            QMessageBox.information(self, "成功", "Excel 参数加载成功！")
        except Exception as e:
            self.log_text.append(f">>> 错误: {str(e)}")
            QMessageBox.critical(self, "错误", f"Excel 读取失败:\n{str(e)}")

    def load_excel(self):
        """读取Excel文件并将数值同步到UI界面 """
        if not self.excel_path:
            QMessageBox.warning(self, "警告", "请先选择 Excel 文件")
            return

        try:
            self.log_text.append(f">>> 正在同步 Excel 数据: {Path(self.excel_path).name}...")
            
            # 调用现有的解析器获取参数对象
            from main import ExcelParser
            parser = ExcelParser(self.excel_path)
            p = parser.parse()

            # 1. 同步几何参数 (Sheet: Geometry)
            g = p.geometry
            self.geom_inputs['L'].setValue(g.L)
            self.geom_inputs['H'].setValue(g.H)
            self.geom_inputs['Tw'].setValue(g.Tw)
            self.geom_inputs['h_pre'].setValue(g.h_pre)
            if 't_cast_cap' in self.geom_inputs:
                self.geom_inputs['t_cast_cap'].setValue(float(getattr(g, 't_cast_cap', 0.0) or 0.0))
            self.geom_inputs['bf_lu'].setValue(g.bf_lu)
            self.geom_inputs['tf_lu'].setValue(g.tf_lu)
            self.geom_inputs['bf_ru'].setValue(g.bf_ru)
            self.geom_inputs['tf_ru'].setValue(g.tf_ru)
            self.geom_inputs['bf_ll'].setValue(g.bf_ll)
            self.geom_inputs['tf_ll'].setValue(g.tf_ll)
            self.geom_inputs['bf_rl'].setValue(g.bf_rl)
            self.geom_inputs['tf_rl'].setValue(g.tf_rl)

            # 2. 同步纵向配筋 (取典型值)
            lr = p.long_rebar
            if lr.left_support_top_A:
                self.rebar_inputs['top_dia'].setValue(lr.left_support_top_A.diameter)
                self.rebar_inputs['top_num'].setValue(lr.left_support_top_A.count)
            if lr.bottom_through_A:
                self.rebar_inputs['bottom_dia'].setValue(lr.bottom_through_A.diameter)
                self.rebar_inputs['bottom_num'].setValue(lr.bottom_through_A.count)

            # 3. 同步箍筋 (Sheet: Stirrups)
            st = p.stirrup
            self.stirrup_inputs['stirrup_dia'].setValue(st.dense_diameter)
            self.stirrup_inputs['stirrup_dense_spacing'].setValue(st.dense_spacing)
            self.stirrup_inputs['stirrup_normal_spacing'].setValue(st.normal_spacing)
            self.stirrup_inputs['stirrup_dense_length'].setValue(st.dense_zone_length)
            self.stirrup_inputs['stirrup_legs'].setValue(st.dense_legs)

            # 4. 同步洞口数据 (仅取第一个洞口作为展示)
            if p.holes:
                h = p.holes[0]
                self.hole_inputs['hole1_x'].setValue(h.x)
                self.hole_inputs['hole1_z'].setValue(h.z)
                self.hole_inputs['hole1_width'].setValue(h.width)
                self.hole_inputs['hole1_height'].setValue(h.height)
                self.hole_inputs['fillet_enabled'].setCurrentText("启用" if h.fillet_radius > 0 else "禁用")
                self.hole_inputs['fillet_radius'].setValue(h.fillet_radius)

            # 5. 同步预应力 (Sheet: Prestress)
            if p.prestress:
                ps = p.prestress
                self.prestress_inputs['enabled'].setCurrentText("启用" if ps.enabled else "禁用")
                self.prestress_inputs['force'].setValue(ps.force)
                try:
                    m = str(getattr(ps, "method", "post_tension") or "post_tension").strip().lower()
                except Exception:
                    m = "post_tension"
                if m == "pretension":
                    self.prestress_inputs['method'].setCurrentText("先张法(pretension)")
                else:
                    self.prestress_inputs['method'].setCurrentText("后张法(post_tension)")
                self.prestress_inputs['duct_diameter'].setValue(ps.duct_diameter)

            self.log_text.append(">>> ✅ Excel 数值已成功同步至 UI 界面！")
            QMessageBox.information(self, "同步成功", "Excel 数据已完美加载到界面，您可以继续微调参数。")
            
        except Exception as e:
            self.log_text.append(f">>> ❌ 同步失败: {str(e)}")
            QMessageBox.critical(self, "同步错误", f"Excel 数据与界面不匹配:\n{str(e)}")


    def _save_ui_params_to_excel(self, excel_path="temp_ui_params.xlsx"):
        """将UI参数保存为Excel文件（100%匹配excel_parser.py的V3.0格式）"""
        wb = Workbook()

        # ========== Sheet 1: Geometry ==========
        ws_geom = wb.active
        ws_geom.title = "Geometry"
        ws_geom.append(["L", "H", "Tw", "bf_lu", "tf_lu", "bf_ru", "tf_ru",
                        "bf_ll", "tf_ll", "bf_rl", "tf_rl", "h_pre", "t_cast_cap"])
        ws_geom.append([
            self.geom_inputs['L'].value(),
            self.geom_inputs['H'].value(),
            self.geom_inputs['Tw'].value(),
            self.geom_inputs['bf_lu'].value(),
            self.geom_inputs['tf_lu'].value(),
            self.geom_inputs['bf_ru'].value(),
            self.geom_inputs['tf_ru'].value(),
            self.geom_inputs['bf_ll'].value(),
            self.geom_inputs['tf_ll'].value(),
            self.geom_inputs['bf_rl'].value(),
            self.geom_inputs['tf_rl'].value(),
            self.geom_inputs['h_pre'].value(),
            self.geom_inputs['t_cast_cap'].value() if ('t_cast_cap' in self.geom_inputs) else 0.0
        ])

        # ========== Sheet 2: Longitudinal Rebar ==========
        ws_rebar = wb.create_sheet("Longitudinal Rebar")
        ws_rebar.append(["Position", "Diameter_A", "Count_A", "Diameter_B", "Count_B", "Extend_Length"])

        # 从UI获取钢筋参数
        top_dia = self.rebar_inputs['top_dia'].value()
        top_num = self.rebar_inputs['top_num'].value()
        bottom_dia = self.rebar_inputs['bottom_dia'].value()
        bottom_num = self.rebar_inputs['bottom_num'].value()

        # 生成4个位置的数据
        ws_rebar.append(["Left Support Top", top_dia, top_num, 0, 0, 500])
        ws_rebar.append(["Mid Span Top", 20, max(2, top_num // 2), 0, 0, 0])
        ws_rebar.append(["Right Support Top", top_dia, top_num, 0, 0, 500])
        ws_rebar.append(["Bottom Through", bottom_dia, bottom_num, 0, 0, 0])

        # ========== Sheet 3: Stirrups ==========
        ws_stirrup = wb.create_sheet("Stirrups")
        ws_stirrup.append(["Zone", "Spacing", "Legs", "Diameter", "Length", "Cover"])

        stirrup_dia = self.stirrup_inputs['stirrup_dia'].value()
        dense_spacing = self.stirrup_inputs['stirrup_dense_spacing'].value()
        normal_spacing = self.stirrup_inputs['stirrup_normal_spacing'].value()
        dense_length = self.stirrup_inputs['stirrup_dense_length'].value()

        ws_stirrup.append(["Dense", dense_spacing, 4, stirrup_dia, dense_length, 25])
        ws_stirrup.append(["Normal", normal_spacing, 2, stirrup_dia, 0, 25])

        # ========== Sheet 4: Holes ==========
        # 注意：Z坐标不是Y！X是纵向位置，Z是竖向位置
        ws_holes = wb.create_sheet("Holes")
        ws_holes.append([
            "X", "Z", "Width", "Height", "Fillet_Radius",
            "SmallBeam_Long_Diameter", "SmallBeam_Long_Count",
            "SmallBeam_Stirrup_Diameter", "SmallBeam_Stirrup_Spacing",
            "Left_Reinf_Length", "Right_Reinf_Length",
            "Side_Stirrup_Spacing", "Side_Stirrup_Diameter", "Side_Stirrup_Legs",
            "Reinf_Extend_Length"
        ])
        
        # 获取UI中的洞口1数据
        hx = self.hole_inputs['hole1_x'].value()
        hz = self.hole_inputs['hole1_z'].value()
        hw = self.hole_inputs['hole1_width'].value()
        hh = self.hole_inputs['hole1_height'].value()

        # 处理倒角逻辑
        fr = 0.0
        if self.hole_inputs['fillet_enabled'].currentText() == "启用":
            fr = self.hole_inputs['fillet_radius'].value()

        # 获取小梁配筋参数 (从UI读取)
        sb_long_dia = self.hole_inputs.get('smallbeam_long_dia', None)
        sb_long_dia = sb_long_dia.value() if sb_long_dia else 16
        sb_long_count = self.hole_inputs.get('smallbeam_long_count', None)
        sb_long_count = sb_long_count.value() if sb_long_count else 2
        sb_stirrup_dia = self.hole_inputs.get('smallbeam_stirrup_dia', None)
        sb_stirrup_dia = sb_stirrup_dia.value() if sb_stirrup_dia else 8
        sb_stirrup_spacing = self.hole_inputs.get('smallbeam_stirrup_spacing', None)
        sb_stirrup_spacing = sb_stirrup_spacing.value() if sb_stirrup_spacing else 150

        # 获取侧边补强参数 (从UI读取)
        left_reinf = self.hole_inputs.get('left_reinf_length', None)
        left_reinf = left_reinf.value() if left_reinf else 500
        right_reinf = self.hole_inputs.get('right_reinf_length', None)
        right_reinf = right_reinf.value() if right_reinf else 500
        side_spacing = self.hole_inputs.get('side_stirrup_spacing', None)
        side_spacing = side_spacing.value() if side_spacing else 100
        side_dia = self.hole_inputs.get('side_stirrup_dia', None)
        side_dia = side_dia.value() if side_dia else 10
        side_legs = self.hole_inputs.get('side_stirrup_legs', None)
        side_legs = side_legs.value() if side_legs else 2
        reinf_extend = self.hole_inputs.get('reinf_extend_length', None)
        reinf_extend = reinf_extend.value() if reinf_extend else 300

        # 写入洞口数据（从UI读取所有参数）
        ws_holes.append([
            hx, hz, hw, hh, fr,
            sb_long_dia, sb_long_count, sb_stirrup_dia, sb_stirrup_spacing,
            left_reinf, right_reinf, side_spacing, side_dia, side_legs, reinf_extend
        ])

        # ========== Sheet 5: Loads ==========
        # 关键：必须有X, X1, X2三个列！
        # Concentrated荷载用X，Distributed荷载用X1和X2
        ws_loads = wb.create_sheet("Loads")
        ws_loads.append(["Case", "Stage", "Type", "X", "X1", "X2", "Direction", "Magnitude"])

        # 获取梁长
        beam_length = self.geom_inputs['L'].value()

        # 获取荷载值 (从UI读取，负值表示向下)
        dead_load_val = self.load_inputs.get('dead_load', None)
        dead_load = -abs(dead_load_val.value()) if dead_load_val else -15.0
        live_load_val = self.load_inputs.get('live_load', None)
        live_load = -abs(live_load_val.value()) if live_load_val else -20.0

        # 施工阶段：仅自重（分布荷载，全跨）
        ws_loads.append(["Dead Load", "Construction", "Distributed", None, 0, beam_length, "Z", dead_load])

        # 使用阶段：自重+活载（分布荷载，全跨）
        ws_loads.append(["Dead Load", "Service", "Distributed", None, 0, beam_length, "Z", dead_load])
        ws_loads.append(["Live Load", "Service", "Distributed", None, 0, beam_length, "Z", live_load])

        # ========== Sheet 6: Prestress ==========
        # 关键：纵向Parameter-Value格式，不是横向列表！
        ws_prestress = wb.create_sheet("Prestress")
        ws_prestress.append(["Parameter", "Value"])

        # 从UI获取预应力参数
        prestress_enabled = (self.prestress_inputs['enabled'].currentText() == "启用")
        prestress_force = self.prestress_inputs['force'].value()
        duct_diameter = self.prestress_inputs['duct_diameter'].value()
        method_text = str(self.prestress_inputs.get('method').currentText() if self.prestress_inputs.get('method') else "后张法(post_tension)")
        prestress_method = "pretension" if ("pretension" in method_text) else "post_tension"
        # 先张法：不挖孔道，写入时强制 duct_diameter=0（避免误配置）
        if prestress_method == "pretension":
            duct_diameter = 0.0

        # 纵向写入参数
        ws_prestress.append(["Enabled", str(prestress_enabled)])
        ws_prestress.append(["Method", prestress_method])
        ws_prestress.append(["Force", prestress_force if prestress_enabled else 0])
        ws_prestress.append(["Duct_Diameter", duct_diameter if prestress_enabled else 0])
        ws_prestress.append(["Path_Type", "straight"])

        # ========== Sheet 7: Boundary ==========
        ws_boundary = wb.create_sheet("Boundary")
        ws_boundary.append(["End", "Dx", "Dy", "Dz", "Rx", "Ry", "Rz", "N", "Vy", "Vz", "Mx", "My", "Mz"])
        ws_boundary.append(["Left", "Fixed", "Fixed", "Fixed", "Free", "Free", "Free", 0, 0, 0, 0, 0, 0])
        ws_boundary.append(["Right", "Free", "Fixed", "Fixed", "Free", "Free", "Free", 0, 0, 0, 0, 0, 0])

        wb.save(excel_path)
        return excel_path

    def generate_model(self):
        """生成模型 - 真实引擎调用"""
        if not ENGINE_AVAILABLE:
            QMessageBox.critical(self, "错误", "主引擎模块未加载！\n请确保 main.py 和 core/ 模块完整。")
            return

        try:
            self.log_text.clear()
            self.log_text.append("="*50)
            self.log_text.append(">>> 开始生成 PKPM-CAE 叠合梁模型...")
            self.log_text.append("="*50)

            # 保存UI参数到临时Excel
            temp_excel = "temp_ui_params.xlsx"
            self.log_text.append(f">>> 正在保存UI参数到 {temp_excel}...")
            self._save_ui_params_to_excel(temp_excel)
            self.log_text.append(">>> ✓ 参数已保存")

            # 创建并启动后台线程
            self.log_text.append(">>> 启动模型生成引擎...")
            self.generation_thread = ModelGenerationThread(temp_excel)
            self.generation_thread.progress.connect(self._on_generation_progress)
            self.generation_thread.finished.connect(self._on_generation_finished)
            self.generation_thread.start()

        except Exception as e:
            import traceback
            self.log_text.append(f">>> ❌ 错误: {str(e)}")
            self.log_text.append(traceback.format_exc())
            QMessageBox.critical(self, "错误", f"模型生成失败:\n{str(e)}")

    def _on_generation_progress(self, message):
        """处理生成进度更新"""
        self.log_text.append(f">>> {message}")

    def _on_generation_finished(self, success, message):
        """处理生成完成"""
        import os

        self.log_text.append("="*50)
        self.log_text.append(f">>> {message}")
        self.log_text.append("="*50)

        # 清理临时文件
        temp_excel = "temp_ui_params.xlsx"
        if os.path.exists(temp_excel):
            try:
                os.remove(temp_excel)
                self.log_text.append(f">>> ✓ 已清理临时文件: {temp_excel}")
            except Exception as e:
                self.log_text.append(f">>> ⚠ 临时文件清理失败: {e}")

        if success:
            QMessageBox.information(self, "成功", message)
        else:
            QMessageBox.critical(self, "错误", message)


def main():
    """主函数"""
    app = QApplication(sys.argv)

    # 设置应用程序字体
    app_font = QFont("Microsoft YaHei", 10)
    app.setFont(app_font)

    window = CompositeBeamUI()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
